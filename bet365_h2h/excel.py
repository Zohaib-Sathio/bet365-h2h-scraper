"""Export the scraped H2H rows to Excel and CSV.

Two layouts are available:

* **simple** — exactly the four columns the brief asks for:
  ``Match | Home Team Win % | Draw % | Away Team Win %``
* **detailed** — the same four, plus league, kick-off and the raw
  head-to-head record the percentages are derived from.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import H2HRow

log = logging.getLogger(__name__)

SIMPLE_COLUMNS = [
    "Match",
    "Home Team Win %",
    "Draw %",
    "Away Team Win %",
]

DETAIL_COLUMNS = [
    "Sport",
    "Continent",
    "Country",
    "League",
    "Date (UTC)",
    "Kick-off (UTC)",
    "Home Team",
    "Away Team",
    "Meetings",
    "Home Wins",
    "Draws",
    "Away Wins",
    "Source (bet365 stats page)",
]

SIMPLE_WIDTHS = [46, 17, 12, 17]
DETAIL_WIDTHS = [12, 14, 18, 30, 12, 14, 22, 22, 11, 12, 9, 12, 46]

HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
LEAGUE_FILL = PatternFill("solid", fgColor="EAF1EE")


def to_records(rows: list[H2HRow], detailed: bool = True) -> list[dict]:
    records = []
    for r in rows:
        m = r.match
        record = {
            "Match": f"{m.home_name} vs {m.away_name}",
            "Home Team Win %": r.home_pct,
            "Draw %": r.draw_pct,
            "Away Team Win %": r.away_pct,
        }
        if detailed:
            record.update(
                {
                    "Sport": m.league.sport,
                    "Continent": m.league.continent,
                    "Country": m.league.country,
                    "League": m.league.name,
                    "Date (UTC)": m.kickoff.strftime("%Y-%m-%d"),
                    "Kick-off (UTC)": m.kickoff.strftime("%H:%M"),
                    "Home Team": m.home_name,
                    "Away Team": m.away_name,
                    "Meetings": r.total_meetings,
                    "Home Wins": r.home_wins,
                    "Draws": r.draws,
                    "Away Wins": r.away_wins,
                    "Source (bet365 stats page)": m.stats_url,
                }
            )
        records.append(record)
    return records


def columns_for(detailed: bool) -> list[str]:
    return SIMPLE_COLUMNS + DETAIL_COLUMNS if detailed else SIMPLE_COLUMNS


def to_dataframe(rows: list[H2HRow], detailed: bool = True) -> pd.DataFrame:
    return pd.DataFrame(to_records(rows, detailed), columns=columns_for(detailed))


def _style(worksheet, n_rows: int, detailed: bool) -> None:
    widths = SIMPLE_WIDTHS + DETAIL_WIDTHS if detailed else SIMPLE_WIDTHS
    for idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = width

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    n_cols = len(widths)
    for row in worksheet.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row[1:4]:  # the three percentage columns
            cell.alignment = Alignment(horizontal="center")

    if detailed:
        # Make the source column clickable so any row can be checked against
        # bet365's own statistics page in one click.
        source_col = n_cols
        for excel_row in range(2, n_rows + 2):
            cell = worksheet.cell(row=excel_row, column=source_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = Font(color="1155CC", underline="single")

        # Banding per league so the sheet reads like the brief's grouped example.
        league_col = len(SIMPLE_COLUMNS) + DETAIL_COLUMNS.index("League") + 1
        previous, shade = None, False
        for excel_row in range(2, n_rows + 2):
            league = worksheet.cell(row=excel_row, column=league_col).value
            if league != previous:
                shade = not shade
                previous = league
            if shade:
                for col in range(1, n_cols + 1):
                    worksheet.cell(row=excel_row, column=col).fill = LEAGUE_FILL

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"


def write_workbook(rows: list[H2HRow], path_or_buffer, detailed: bool = True) -> None:
    """Write the styled workbook to a path or an open binary buffer."""
    frame = to_dataframe(rows, detailed)
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="H2H Percentages", index=False)
        _style(writer.sheets["H2H Percentages"], len(frame), detailed)


def export(
    rows: list[H2HRow], out_dir: Path, stem: str, detailed: bool = True
) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    frame = to_dataframe(rows, detailed)

    csv_path = out_dir / f"{stem}.csv"
    frame.to_csv(csv_path, index=False, encoding="utf-8-sig")

    xlsx_path = out_dir / f"{stem}.xlsx"
    write_workbook(rows, xlsx_path, detailed)

    log.info("wrote %s (%d rows)", xlsx_path.name, len(frame))
    return xlsx_path, csv_path
